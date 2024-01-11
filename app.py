import time
import io
from io import StringIO

import pandas as pd
import numpy as np

import xlsxwriter
import openpyxl
from openpyxl import load_workbook

import streamlit as st
import plotly_express as px
from streamlit_option_menu import option_menu

from openai import OpenAI

try:
    from config import *
    client = OpenAI(api_key=API_KEY)
except:
    client = OpenAI(api_key=st.secrets["api_key"])

import pandas as pd
import sys
import os 

# Add the project root directory to the Python path
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
sys.path.append(project_root)

from src.embedding.text_embedder import TextProcessor

# instance of text processor
tp = TextProcessor()

# try:
#     client = OpenAI(api_key=API_KEY)
# except:
#     client = OpenAI(api_key=st.secrets["api_key"])


# st.set_page_config(layout='wide')

@st.cache_data
def load_data(path):
    if '.parquet' in path:
        return pd.read_parquet(path)
    if '.csv' in path:
        return pd.read_csv(path)


def concat_df(list_of_df):
    """
    This function is for the purpose of combining grant programs 
    from different agencies into one dataframe.
    Further processing is necessary and is accomplished with the 
    "process_df" function.
    """
    grants = pd.concat(list_of_df).reset_index(drop=True)
    return grants


def process_df(df):
    df = df.rename(columns={'parent_name': 'agency'})
    df = df.rename(columns={'sbir_topic_link': 'url'})
    df = df.rename(columns={'topic_title': 'title'})
    df = df.rename(columns={'opportunity_number': 'topic_number'})
    df = df.query('text.str.len() > 250')
    df = df.reset_index(drop=True)
    return df


def get_embedding(text, model="text-embedding-ada-002"):
    return client.embeddings.create(input = [text], model=model).data[0].embedding


def clear_query():
    st.session_state['query_text'] = ""
    st.session_state['text_box_button'] = False

    
def convert_df_to_excel(df):
    """
    This function converts a dataframe into a format that is exportable 
    via a download button. It also merges cells based on the 'company' column.
    """
    # First, save the DataFrame to a BytesIO buffer using Pandas
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    
    # Load the workbook and modify it using openpyxl
    output.seek(0) # Rewind the buffer
    wb = load_workbook(output)
    ws = wb.active

    # Apply merge to each column independently
    for col in ['A', 'B', 'C']:
        for range_info in find_rows_to_merge(ws, col):
            ws.merge_cells(f'{col}{range_info["start_row"]}:{col}{range_info["end_row"]}')

    # Save the modified workbook to a new BytesIO buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()
    

def convert_df_to_csv(df):
    """
    This function converts a DataFrame into a CSV format that is exportable 
    via a download button.
    """
    # Create a StringIO buffer
    output = StringIO()
    
    # Write the DataFrame to the buffer as a CSV
    df.to_csv(output, index=False)

    # Return the buffer's content in a format suitable for streamlit.download_button
    return output.getvalue()


def read_file(uploaded_file):
    file_name = uploaded_file.name
    if '.csv' in file_name:
        df = pd.read_csv(uploaded_file, encoding='utf-8')
    if '.xlsx' in file_name:
        df = pd.read_excel(uploaded_file)
    return df


def find_rows_to_merge(ws, column):
    merge_ranges = []
    start_row = 1
    current_value = None

    for row in range(2, ws.max_row + 1):  # Starting from 2 to skip header row
        cell_value = ws[f'{column}{row}'].value
        if cell_value != current_value:
            end_row = row - 1
            if end_row > start_row:
                merge_ranges.append({'start_row': start_row, 'end_row': end_row})
            start_row = row
            current_value = cell_value

    # Add the last range
    if ws.max_row > start_row:
        merge_ranges.append({'start_row': start_row, 'end_row': ws.max_row})

    return merge_ranges


def modified_jaccard(set1, set2):
    # Convert lists to sets (if they aren't sets already)
    set1 = set(set1)
    set2 = set(set2)

    # Calculate intersection and union
    intersection = set1.intersection(set2)

    # Compute Jaccard Index
    jaccard_index = len(intersection) / len(set2)
    return jaccard_index


# load data
dod = load_data('data/dod_processed.parquet')
doe = load_data('data/doe_processed.parquet')
hhs = load_data('data/hhs_processed.parquet')

# process data
dod = process_df(dod)
doe = process_df(doe)
hhs = process_df(hhs)

grants = concat_df([dod, doe, hhs])



if "button1" not in st.session_state:
    st.session_state["button1"] = False

if "button2" not in st.session_state:
    st.session_state["button2"] = False

if "grant_recommendations" not in st.session_state:
    st.session_state["grant_recommendations"] = pd.DataFrame()

if "text_box_button" not in st.session_state:
    st.session_state['text_box_button'] = False

st.session_state["download_csv"] = True

# create sidebar menu with options
selected = option_menu(
    menu_title=None,
    menu_icon='cast',
    default_index=0,
    options=['Drag & Drop', 'Text Box'],
    orientation='horizontal',
    icons=['droplet', 'card-text'],
    styles= {'container': {
                'font-size': '12px'
    }}
)


if selected == 'Drag & Drop':

    # Upload csv
    uploaded_file  = st.file_uploader('Choose a file', accept_multiple_files=False)

    if uploaded_file is not None:
        dataframe = read_file(uploaded_file)

        # drop duplicate abstracts
        dataframe = (dataframe
                     .assign(temp_summary = lambda df: df['summary'].str.lower())
                     .drop_duplicates(subset='temp_summary')
                     .drop('temp_summary', axis=1)
                     .reset_index(drop=True)
                     )

    col1, col2, col3, col4 = st.columns([3, 3, 5, 3])

    with col1:
        find_program = st.button('Find Programs')

    # filter_columns = st.columns([5, 7])
    # with filter_columns[0]:
    #     with st.expander('Filters'):
    #         agencies = st.multiselect(label='Choose Agency', options=grants['agency'].unique())

    if find_program:
        st.session_state['button1'] = True
        st.session_state['download_csv'] = False

    if st.session_state['button1'] and uploaded_file is not None and st.session_state['download_csv'] == False:
        column_names = ['company', 'project_title', 'summary', 'agency', 'grant_title', 'url', 'text']
        grant_recommendations = pd.DataFrame(columns=column_names)
        for index, row in dataframe.iterrows():
            company = row['company']
            summary = row['summary']
            title = row['title']
            query_embedding = get_embedding(summary)

            # tokens and ngrams for abstract text
            tokens = tp.generate_tokens(text=summary)
            unigrams = tp.generate_ngrams(tokens, 1)
            bigrams = tp.generate_ngrams(tokens, 2)
            trigrams = tp.generate_ngrams(tokens, 3)
            temp_df = (grants
                      .assign(
                          jaccard_unigram=grants['unigrams'].apply(lambda x: modified_jaccard(x, unigrams)),
                          jaccard_bigram=grants['bigrams'].apply(lambda x: modified_jaccard(x, bigrams)),
                          jaccard_trigram=grants['trigrams'].apply(lambda x: modified_jaccard(x, trigrams)),
                          cosine_similarity=grants['text_embedded'].apply(lambda y: np.dot(query_embedding, y)),
                          description_len=grants['text'].str.len()
                          )
                      .query('cosine_similarity >= 0.80 and jaccard_bigram != 0')
                      .groupby(['topic_number'], as_index=False)
                      .agg(
                          title=('title', 'max'),
                          topic_number=('topic_number', 'max'),
                          text=('text', 'sum'),
                          jaccard_bigram=('jaccard_bigram', 'max'),
                          jaccard_trigram=('jaccard_trigram', 'max'),
                          cosine_similarity=('cosine_similarity', 'max'),
                          description_len=('description_len', 'median'),
                          url=('url', 'max'),
                          agency=('agency', 'max')
                          )
                    ).sort_values(by='cosine_similarity', ascending=False).head(3)
            for index2, row2 in temp_df.iterrows():
                agency = row2['agency']
                grant_title = row2['title']
                url = row2['url']
                text = row2['text']
                cosine_similarity = row2['cosine_similarity']
                new_row = {'company': company, 'project_title': title, 'summary': summary, 'agency': agency, 'grant_title': grant_title, 'text': text, 'url': url}
                new_row_df = pd.DataFrame([new_row])
                grant_recommendations = pd.concat([grant_recommendations, new_row_df], ignore_index=True)
        st.session_state["grant_recommendations"] = grant_recommendations


    if not st.session_state["grant_recommendations"].empty:
        st.divider()
        st.subheader('Grant Recommendations')
        grant_recommendations = st.session_state["grant_recommendations"]
        st.write(grant_recommendations)


        excel_file = convert_df_to_excel(grant_recommendations)
        

        if st.download_button("Press to Download", excel_file, "grant_recommendations.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key='download-excel'):
            st.session_state['download_excel'] = True
        # csv = convert_df_to_csv(grant_recommendations)
        # if st.download_button("Press to Download",csv,"grant_recommendations.csv","text/csv", key='download-csv'):
        #     st.session_state['download_csv'] = True


if selected == 'Text Box':
    # Text Input
    query = st.text_area('Enter description:', max_chars=3000, key='query_text', height=300)

    col1, col2, col3, col4 = st.columns([3, 3, 5, 3])

    with col1:
        find_program = st.button('Find Programs')
    with col2:
        delete_abstract = st.button('Clear Keywords', on_click=clear_query)

    if find_program:
        st.session_state['text_box_button'] = True

    if st.session_state['text_box_button'] & len(query) > 0:
        # tokens and ngrams for abstract text
        query_embedding = get_embedding(query)
        tokens = tp.generate_tokens(text=query)
        unigrams = tp.generate_ngrams(tokens, 1)
        bigrams = tp.generate_ngrams(tokens, 2)
        trigrams = tp.generate_ngrams(tokens, 3)

        temp_df = (grants
                  .assign(
                      jaccard_unigram=grants['unigrams'].apply(lambda x: modified_jaccard(x, unigrams)),
                      jaccard_bigram=grants['bigrams'].apply(lambda x: modified_jaccard(x, bigrams)),
                      jaccard_trigram=grants['trigrams'].apply(lambda x: modified_jaccard(x, trigrams)),
                      cosine_similarity=grants['text_embedded'].apply(lambda y: np.dot(query_embedding, y)),
                      description_len=grants['text'].str.len()
                      )
                  .query('cosine_similarity >= 0.80 and jaccard_bigram != 0')
                  .groupby(['topic_number'], as_index=False)
                  .agg(title=('title', 'max'),
                       topic_number=('topic_number', 'max'),
                       text=('text', 'sum'),
                       jaccard_bigram=('jaccard_bigram', 'max'),
                       jaccard_trigram=('jaccard_trigram', 'max'),
                       cosine_similarity=('cosine_similarity', 'max'),
                       description_len=('description_len', 'median'),
                       url=('url', 'max')
                       )
                  ).sort_values(by='cosine_similarity', ascending=False)
        


        st.write(temp_df[['title', 'topic_number', 'text', 'jaccard_bigram', 'jaccard_trigram', 'description_len', 'cosine_similarity', 'url']])
        fig = px.scatter(temp_df, x='jaccard_bigram', y='cosine_similarity', trendline="ols", trendline_color_override='red', opacity=0.5)
        st.plotly_chart(fig)



    # for index, row in grants.iloc[similarity_index].iterrows():
    #     st.markdown(f'#### {row["title"]}')
    #     with st.expander('Show Text'):
    #         st.write(row['text'])
    #     st.write(row['url'])


        # st.write('Number:', row['number'])
        # if st.session_state['button1']:
        #     if st.button('AI Report', key=f'ai{index}'):
        #         text = row['text']
        #         completion = client.chat.completions.create(
        #         model="gpt-3.5-turbo-1106",
        #         messages=[
        #             {"role": "system", "content": "You aid people in understand grant proposals. Your purpose is to compare the abstract of a proposal to the grant program description. Explain in bulleted format in less than 700 words why the abstract and grant program match or don't match. Be sure to mention specific technologies, innovations, research topics, etc. that are in the provided grant and abstract text."},
        #             {"role": "user", "content": f"Here is my abstract: {query}. Here is the grant description: {text}"}
        #         ]
        #         )
        #         st.text_area(label='', value=f'{completion.choices[0].message.content}', height=700)

        # st.divider()

